import datetime
import random

import openpyxl
# 读取Excel文件
# try:
#     # 打开工作表
#     wb=openpyxl.load_workbook(filename="../files/PythonExcel.xlsx")
#     # 打印工作表名称
#     print(wb.sheetnames)
#     # 获取工作表 --->Worksheet
#     sheet=wb.worksheets[0]
#     # 获取单元格的范围
#     print(sheet.dimensions)
#     # 获得行数和列数
#     print(sheet.max_row,sheet.max_column)
#     # 获取单元格内容
#     print(sheet.cell(2,1).value)
#     print(sheet['C2'].value)
#     # 如果不存在该单元格则返回None
#     print(sheet['C3'].value)
#     # 获取多个单元格
#     print(sheet['A1':'C2'])
#     #读取所有单元格的数据
#     for row_ch in range(1,sheet.max_row):
#         for col_ch in 'ABC':
#             value=sheet[f'{col_ch}{row_ch}'].value
#             if type(value) ==int:
#                 print(f'{value:<10d}',end='\t')
#         print()
# except FileNotFoundError:
#     print("File not found")


#写入Excel文件
try:
    #第一步，创建工作簿
    wb=openpyxl.Workbook()
    #第二部，添加工作表
    sheet=wb.active
    sheet.title='学生信息'
    titles=['姓名','年龄','成绩']
    #第三步，添加标题
    for col_index,title in enumerate(titles):
        sheet.cell(1,col_index+1,value=title)#表中的第一行为标题
    #第四步，添加数据
    names= ['张三','李四','王五']
    for row_index,name in enumerate(names):
        sheet.cell(row_index+2,1,value=name)#表中的第二行开始为数据，第一列为姓名
        for col_index in range(2,len(titles)+1):#表中的第二列开始为数据（即从第二行第二列开始才是数据）
            sheet.cell(row_index+2,col_index,random.randrange(60,100))
    #第五步，保存工作簿
    wb.save(filename="../files/PythonExcel2.xlsx")
except FileNotFoundError:
    print("File not found")